"""Reusable attachment text extraction helpers."""

from __future__ import annotations

import logging
import os
from pathlib import Path
import shutil

try:
    import pytesseract  # type: ignore
except Exception:  # noqa: BLE001
    pytesseract = None  # type: ignore[assignment]

MAX_ATTACHMENT_TEXT = 20_000
MAX_CSV_CHARS = 12_000
MAX_SPREADSHEET_ROWS = 2_000
AUDIO_EXT = {".mp3", ".m4a", ".wav", ".mp4", ".webm"}
PDF_EXT = {".pdf"}
DOC_EXT = {".doc", ".docx"}
EXCEL_EXT = {".xls", ".xlsx"}
TEXT_EXT = {".txt", ".csv"}
SUPPORTED_ATTACHMENT_EXTENSIONS = PDF_EXT | DOC_EXT | EXCEL_EXT | TEXT_EXT | AUDIO_EXT
PDF_OCR_MIN_TEXT_LENGTH = 50
POPPLER_PATH = r"C:\poppler\Library\bin"
MAX_AUDIO_FILE_SIZE_BYTES = 25 * 1024 * 1024

logger = logging.getLogger(__name__)


def configure_tesseract() -> None:
    """Configure pytesseract to find tesseract.exe in common locations."""
    if pytesseract is None:
        logger.warning("pytesseract is not available; OCR fallback may be limited")
        return

    try:
        tesseract_from_path = shutil.which("tesseract")
        if tesseract_from_path:
            logger.info("Tesseract found in PATH: %s", tesseract_from_path)
            return

        default_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(default_path):
            pytesseract.pytesseract.tesseract_cmd = default_path
            logger.info("Tesseract configured manually: %s", default_path)
            return

        logger.warning("Tesseract executable not found on system")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Tesseract configuration failed: %s", exc)


configure_tesseract()


def extract_text_from_attachment(file_path: str, filename: str = "") -> str:
    """Extract plain text from a supported attachment path."""
    path = Path(str(file_path or "").strip())
    suffix = _detect_extension(filename=filename, file_path=str(path))
    if suffix not in SUPPORTED_ATTACHMENT_EXTENSIONS or not path.exists():
        return ""

    try:
        if suffix in AUDIO_EXT:
            return transcribe_audio(str(path))
        if suffix in PDF_EXT:
            return extract_pdf_text(str(path))
        if suffix in DOC_EXT:
            return extract_doc_text(str(path))
        if suffix in EXCEL_EXT:
            return extract_excel_summary(str(path))
        if suffix == ".txt":
            return _extract_txt(path)
        if suffix == ".csv":
            return _extract_csv(path)
    except Exception:  # noqa: BLE001
        logger.exception("Attachment extraction failed: %s", path.name)
        return ""

    return ""


def extract_text_from_attachments(attachments: list[dict[str, str]]) -> str:
    combined_text, _content_types = extract_text_and_types_from_attachments(attachments)
    return combined_text


def extract_text_and_types_from_attachments(attachments: list[dict[str, str]]) -> tuple[str, list[str]]:
    """Extract and combine text from multiple supported attachments."""
    blocks: list[str] = []
    content_types: list[str] = []
    for attachment in attachments or []:
        local_path = str(attachment.get("file_path") or attachment.get("local_path") or "").strip()
        filename = str(attachment.get("filename") or Path(local_path).name or "adjunto").strip() or "adjunto"
        if not local_path:
            continue

        suffix = _detect_extension(filename=filename, file_path=local_path)
        if suffix not in SUPPORTED_ATTACHMENT_EXTENSIONS:
            continue

        logger.info("Extracting text from attachment: %s", filename)
        logger.info("Attachment type supported: %s", suffix.lstrip("."))
        try:
            extracted, content_type = process_attachment(local_path)
        except Exception:  # noqa: BLE001
            logger.exception("Attachment extraction failed: %s", filename)
            continue
        extracted = extracted.strip()
        logger.info("Attachment text extracted: %s characters", len(extracted))
        if not extracted:
            continue
        if content_type not in content_types:
            content_types.append(content_type)

        blocks.append(
            f"ATTACHMENT: {filename}\n"
            "--------------------------------\n"
            f"{extracted}"
        )

    combined = "\n\n".join(blocks).strip()
    if len(combined) > MAX_ATTACHMENT_TEXT:
        return combined[:MAX_ATTACHMENT_TEXT], content_types
    return combined, content_types


def process_attachment(file_path: str) -> tuple[str, str]:
    """Process one attachment and return extracted text plus normalized content type."""
    path = Path(str(file_path or "").strip())
    ext = get_extension(str(path))

    if ext in AUDIO_EXT:
        text = transcribe_audio(str(path))
        return text, "audio_meeting"
    if ext in PDF_EXT:
        return extract_pdf_text(str(path)), "pdf"
    if ext in DOC_EXT:
        return extract_doc_text(str(path)), "doc"
    if ext in EXCEL_EXT:
        return extract_excel_summary(str(path)), "excel"
    if ext == ".txt":
        return _extract_txt(path), "txt"
    if ext == ".csv":
        return _extract_csv(path), "csv"
    raise ValueError("Tipo de adjunto no soportado")


def transcribe_audio(file_path: str) -> str:
    """Transcribe audio attachments and cache transcript alongside the source file."""
    path = Path(str(file_path or "").strip())
    cached_transcript_path = path.with_suffix(".txt")
    if cached_transcript_path.exists():
        logger.info("Using cached audio transcript for %s", path.name)
        return _read_text_with_fallback(cached_transcript_path).strip()

    if path.stat().st_size > MAX_AUDIO_FILE_SIZE_BYTES:
        raise ValueError("Audio demasiado grande")

    from openai import OpenAI

    from app.utils.openai_client import load_api_key

    client = OpenAI(api_key=load_api_key())
    with path.open("rb") as audio_file:
        transcription = client.audio.transcriptions.create(
            model="gpt-4o-mini-transcribe",
            file=audio_file,
            response_format="text",
        )

    transcript_text = str(getattr(transcription, "text", transcription) or "").strip()
    if transcript_text:
        cached_transcript_path.write_text(transcript_text, encoding="utf-8")
    return transcript_text


def get_extension(file_path: str) -> str:
    """Return normalized attachment extension."""
    return Path(file_path or "").suffix.lower().strip()


def extract_pdf_text(file_path: str) -> str:
    return _extract_pdf(Path(file_path))


def extract_doc_text(file_path: str) -> str:
    path = Path(file_path)
    if path.suffix.lower() == ".docx":
        return _extract_docx(path)
    logger.warning("Legacy DOC extraction not available for %s", path.name)
    return ""


def extract_excel_summary(file_path: str) -> str:
    path = Path(file_path)
    if path.suffix.lower() == ".xlsx":
        return _extract_xlsx(path)
    return _extract_xls(path)


def _extract_pdf(path: Path) -> str:
    texts: list[str] = []

    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(str(path)) as doc:
            for page in doc.pages:
                text = str(page.extract_text() or "").strip()
                if text:
                    texts.append(text)
    except Exception:  # noqa: BLE001
        logger.warning("pdfplumber extraction failed for %s; trying PyPDF2 fallback", path.name)

    if texts:
        return "\n\n".join(texts)

    try:
        from PyPDF2 import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        for page in reader.pages:
            text = str(page.extract_text() or "").strip()
            if text:
                texts.append(text)
    except Exception:  # noqa: BLE001
        logger.warning("PyPDF2 extraction failed for %s; trying PyMuPDF fallback", path.name)

    if texts:
        return "\n\n".join(texts)

    try:
        import fitz  # type: ignore

        with fitz.open(path) as doc:
            for page in doc:
                text = (page.get_text("text") or "").strip()
                if text:
                    texts.append(text)
    except Exception:  # noqa: BLE001
        logger.warning("PyMuPDF extraction failed for %s; trying pdfminer fallback", path.name)

    if texts:
        return "\n\n".join(texts)

    try:
        from pdfminer.high_level import extract_text  # type: ignore

        text = str(extract_text(str(path)) or "").strip()
        if text:
            texts.append(text)
    except Exception:  # noqa: BLE001
        logger.warning("pdfminer extraction failed for %s", path.name)

    extracted_text = "\n\n".join(texts).strip()
    if len(extracted_text) >= PDF_OCR_MIN_TEXT_LENGTH:
        return extracted_text

    logger.info("PDF contained no text, running OCR fallback")
    ocr_text = _extract_pdf_with_ocr(path)
    if not ocr_text:
        return extracted_text
    if extracted_text:
        return f"{extracted_text}\n\n{ocr_text}".strip()
    return ocr_text


def _extract_pdf_with_ocr(path: Path) -> str:
    try:
        from pdf2image import convert_from_bytes  # type: ignore
    except Exception:  # noqa: BLE001
        logger.warning("OCR dependencies are not available for %s", path.name)
        return ""

    if pytesseract is None:
        logger.warning("OCR dependencies are not available for %s", path.name)
        return ""

    logger.info("OCR fallback started")

    try:
        pdf_bytes = path.read_bytes()
        images = convert_from_bytes(pdf_bytes, dpi=300, poppler_path=POPPLER_PATH)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to convert PDF to images for OCR: %s", exc)
        return ""

    texts: list[str] = []
    for image in images:
        try:
            text = str(pytesseract.image_to_string(image) or "").strip()
        except Exception:  # noqa: BLE001
            logger.warning("OCR page extraction failed for %s", path.name)
            continue
        if text:
            texts.append(text)
    extracted = "\n\n".join(texts).strip()
    logger.info("OCR extracted %s characters", len(extracted))
    return extracted


def _detect_extension(filename: str, file_path: str) -> str:
    candidates = [str(filename or "").strip().lower(), Path(file_path or "").name.lower()]
    for candidate in candidates:
        if candidate.endswith(".pdf"):
            return ".pdf"
        if candidate.endswith(".xlsx"):
            return ".xlsx"
        if candidate.endswith(".xls"):
            return ".xls"
        if candidate.endswith(".docx"):
            return ".docx"
        if candidate.endswith(".doc"):
            return ".doc"
        if candidate.endswith(".txt"):
            return ".txt"
        if candidate.endswith(".csv"):
            return ".csv"
        if candidate.endswith(".mp3"):
            return ".mp3"
        if candidate.endswith(".m4a"):
            return ".m4a"
        if candidate.endswith(".wav"):
            return ".wav"
        if candidate.endswith(".mp4"):
            return ".mp4"
        if candidate.endswith(".webm"):
            return ".webm"
    return ""


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document  # type: ignore
    except Exception:  # noqa: BLE001
        logger.warning("python-docx is not available; cannot read %s", path.name)
        return ""

    document = Document(str(path))
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text and paragraph.text.strip()]
    return "\n".join(paragraphs)


def _extract_txt(path: Path) -> str:
    return _read_text_with_fallback(path).strip()


def _extract_csv(path: Path) -> str:
    content = _read_text_with_fallback(path)
    if len(content) > MAX_CSV_CHARS:
        return content[:MAX_CSV_CHARS]
    return content


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:  # noqa: BLE001
        logger.warning("openpyxl is not available; cannot read %s", path.name)
        return ""

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    blocks: list[str] = []
    rows_read = 0

    for sheet in workbook.worksheets:
        sheet_lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            if rows_read >= MAX_SPREADSHEET_ROWS:
                break
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            line = " | ".join(value for value in values if value)
            if line:
                sheet_lines.append(line)
            rows_read += 1
        if sheet_lines:
            blocks.append(f"[{sheet.title}]\n" + "\n".join(sheet_lines))
        if rows_read >= MAX_SPREADSHEET_ROWS:
            break

    workbook.close()
    return "\n\n".join(blocks).strip()


def _extract_xls(path: Path) -> str:
    try:
        import xlrd  # type: ignore
    except Exception:  # noqa: BLE001
        logger.warning("xlrd is not available; cannot read %s", path.name)
        return ""

    try:
        workbook = xlrd.open_workbook(str(path))
    except Exception:  # noqa: BLE001
        logger.warning("Failed to parse spreadsheet %s", path.name)
        return ""

    blocks: list[str] = []
    rows_read = 0
    for sheet in workbook.sheets():
        sheet_lines: list[str] = []
        for row_index in range(sheet.nrows):
            if rows_read >= MAX_SPREADSHEET_ROWS:
                break
            values = [str(cell).strip() for cell in sheet.row_values(row_index)]
            line = " | ".join(value for value in values if value)
            if line:
                sheet_lines.append(line)
            rows_read += 1
        if sheet_lines:
            blocks.append(f"[{sheet.name}]\n" + "\n".join(sheet_lines))
        if rows_read >= MAX_SPREADSHEET_ROWS:
            break

    return "\n\n".join(blocks).strip()


def _read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")
